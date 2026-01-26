#!/usr/bin/env python3
"""
Excel to Energistics UOM Dictionary XML Converter

Reads a populated UOM Dictionary Excel workbook and generates
an Energistics-compliant XML UOM dictionary file.

Usage:
    python excel_to_uom_xml.py <input_xlsx> <output_xml> [namespace]
    
Example:
    python excel_to_uom_xml.py uom_output.xlsx uom_dictionary.xml http://www.energistics.org/energyml/data/uomv1
"""

import sys
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from xml.dom import minidom


def read_excel_data(excel_file):
    """Read data from the populated Excel workbook."""
    
    wb = load_workbook(excel_file, data_only=True)
    
    data = {
        'version': '',
        'namespace': '',
        'schemaLocation': '',
        'title': '',
        'originator': '',
        'description': '',
        'unit_dimensions': [],
        'quantity_classes': [],
        'units': [],
        'references': [],
        'prefixes': []
    }
    
    # Read Metadata
    if 'Metadata' in wb.sheetnames:
        metadata = wb['Metadata']
        data['version'] = str(metadata['B1'].value or '')
        # Read all metadata fields
        for row in range(3, 15):
            label = metadata.cell(row, 1).value
            value = metadata.cell(row, 2).value
            if label and value:
                label_lower = str(label).lower()
                if 'namespace' in label_lower:
                    data['namespace'] = str(value)
                elif 'schema' in label_lower and 'location' in label_lower:
                    data['schemaLocation'] = str(value)
                elif 'title' in label_lower:
                    data['title'] = str(value)
                elif 'originator' in label_lower:
                    data['originator'] = str(value)
                elif 'description' in label_lower:
                    data['description'] = str(value)
    
    # Read Unit Dimensions
    if 'Unit Dimensions' in wb.sheetnames:
        sheet = wb['Unit Dimensions']
        for row in range(2, sheet.max_row + 1):
            name = sheet.cell(row, 1).value
            if name:  # Only process rows with data
                data['unit_dimensions'].append({
                    'name': str(name),
                    'dimension': str(sheet.cell(row, 2).value or ''),
                    'baseForConversion': str(sheet.cell(row, 3).value or ''),
                    'canonicalUnit': str(sheet.cell(row, 4).value or ''),
                    'description': str(sheet.cell(row, 5).value or '')
                })
    
    # Read Quantity Classes
    if 'Quantity Classes' in wb.sheetnames:
        sheet = wb['Quantity Classes']
        for row in range(2, sheet.max_row + 1):
            name = sheet.cell(row, 1).value
            if name:
                member_units_str = str(sheet.cell(row, 6).value or '')
                member_units = [mu.strip() for mu in member_units_str.split(',') if mu.strip()]
                
                data['quantity_classes'].append({
                    'name': str(name),
                    'dimension': str(sheet.cell(row, 2).value or ''),
                    'baseForConversion': str(sheet.cell(row, 3).value or ''),
                    'alternativeBase': str(sheet.cell(row, 4).value or ''),
                    'description': str(sheet.cell(row, 5).value or ''),
                    'memberUnits': member_units
                })
    
    # Read Units
    if 'Units' in wb.sheetnames:
        sheet = wb['Units']
        for row in range(3, sheet.max_row + 1):  # Skip note row
            symbol = sheet.cell(row, 1).value
            if symbol:
                is_base_val = str(sheet.cell(row, 6).value or '').upper()
                is_base = is_base_val in ('TRUE', '1', 'YES')
                
                data['units'].append({
                    'symbol': str(symbol),
                    'name': str(sheet.cell(row, 2).value or ''),
                    'dimension': str(sheet.cell(row, 3).value or ''),
                    'isSI': str(sheet.cell(row, 4).value or ''),
                    'category': str(sheet.cell(row, 5).value or ''),
                    'isBase': is_base,
                    'baseUnit': '' if is_base else str(sheet.cell(row, 7).value or ''),
                    'conversionRef': '' if is_base else str(sheet.cell(row, 8).value or ''),
                    'isExact': '' if is_base else str(sheet.cell(row, 9).value or ''),
                    'A': '' if is_base else str(sheet.cell(row, 10).value or ''),
                    'B': '' if is_base else str(sheet.cell(row, 11).value or ''),
                    'C': '' if is_base else str(sheet.cell(row, 12).value or ''),
                    'D': '' if is_base else str(sheet.cell(row, 13).value or ''),
                    'underlyingDef': str(sheet.cell(row, 14).value or ''),
                    'description': str(sheet.cell(row, 15).value or '')
                })
    
    # Read References
    if 'References' in wb.sheetnames:
        sheet = wb['References']
        for row in range(2, sheet.max_row + 1):
            ref_id = sheet.cell(row, 1).value
            if ref_id:
                data['references'].append({
                    'ID': str(ref_id),
                    'description': str(sheet.cell(row, 2).value or '')
                })
    
    # Read Prefixes
    if 'Prefixes' in wb.sheetnames:
        sheet = wb['Prefixes']
        for row in range(2, sheet.max_row + 1):
            symbol = sheet.cell(row, 1).value
            if symbol:
                data['prefixes'].append({
                    'symbol': str(symbol),
                    'name': str(sheet.cell(row, 2).value or ''),
                    'multiplier': str(sheet.cell(row, 3).value or ''),
                    'commonName': str(sheet.cell(row, 4).value or '')
                })
    
    return data


def create_xml(data, namespace=None):
    """Create XML structure from data."""
    
    # Use namespace from data if available, otherwise use provided or default
    if not namespace:
        namespace = data.get('namespace', 'http://www.energistics.org/energyml/data/uomv1')
    
    # Register namespaces
    ET.register_namespace('', namespace)
    ET.register_namespace('xsi', 'http://www.w3.org/2001/XMLSchema-instance')
    
    xsi_ns = 'http://www.w3.org/2001/XMLSchema-instance'
    
    # Create root element with all attributes
    root_attrib = {'version': data['version']}
    
    # Add schema location if available
    if data.get('schemaLocation'):
        root_attrib[f'{{{xsi_ns}}}schemaLocation'] = data['schemaLocation']
    
    root = ET.Element(f'{{{namespace}}}uomDictionary', attrib=root_attrib)
    
    # Add citation information
    if data.get('title'):
        title_elem = ET.SubElement(root, f'{{{namespace}}}title')
        title_elem.text = data['title']
    
    if data.get('originator'):
        orig_elem = ET.SubElement(root, f'{{{namespace}}}originator')
        orig_elem.text = data['originator']
    
    if data.get('description'):
        desc_elem = ET.SubElement(root, f'{{{namespace}}}description')
        desc_elem.text = data['description']
    
    # Unit Dimension Set
    unit_dim_set = ET.SubElement(root, f'{{{namespace}}}unitDimensionSet', version=data['version'])
    for ud in data['unit_dimensions']:
        unit_dim = ET.SubElement(unit_dim_set, f'{{{namespace}}}unitDimension')
        
        name = ET.SubElement(unit_dim, f'{{{namespace}}}name')
        name.text = ud['name']
        
        dimension = ET.SubElement(unit_dim, f'{{{namespace}}}dimension')
        dimension.text = ud['dimension']
        
        base_conv = ET.SubElement(unit_dim, f'{{{namespace}}}baseForConversion')
        base_conv.text = ud['baseForConversion']
        
        canonical = ET.SubElement(unit_dim, f'{{{namespace}}}canonicalUnit')
        canonical.text = ud['canonicalUnit']
        
        if ud['description']:
            desc = ET.SubElement(unit_dim, f'{{{namespace}}}description')
            desc.text = ud['description']
    
    # Quantity Class Set
    qty_class_set = ET.SubElement(root, f'{{{namespace}}}quantityClassSet', version=data['version'])
    for qc in data['quantity_classes']:
        qty_class = ET.SubElement(qty_class_set, f'{{{namespace}}}quantityClass')
        
        name = ET.SubElement(qty_class, f'{{{namespace}}}name')
        name.text = qc['name']
        
        dimension = ET.SubElement(qty_class, f'{{{namespace}}}dimension')
        dimension.text = qc['dimension']
        
        base_conv = ET.SubElement(qty_class, f'{{{namespace}}}baseForConversion')
        base_conv.text = qc['baseForConversion']
        
        if qc['alternativeBase']:
            alt_base = ET.SubElement(qty_class, f'{{{namespace}}}alternativeBase')
            alt_base.text = qc['alternativeBase']
        
        if qc['description']:
            desc = ET.SubElement(qty_class, f'{{{namespace}}}description')
            desc.text = qc['description']
        
        for member in qc['memberUnits']:
            member_unit = ET.SubElement(qty_class, f'{{{namespace}}}memberUnit')
            member_unit.text = member
    
    # Unit Set
    unit_set = ET.SubElement(root, f'{{{namespace}}}unitSet', version=data['version'])
    for unit in data['units']:
        unit_elem = ET.SubElement(unit_set, f'{{{namespace}}}unit')
        
        symbol = ET.SubElement(unit_elem, f'{{{namespace}}}symbol')
        symbol.text = unit['symbol']
        
        name = ET.SubElement(unit_elem, f'{{{namespace}}}name')
        name.text = unit['name']
        
        dimension = ET.SubElement(unit_elem, f'{{{namespace}}}dimension')
        dimension.text = unit['dimension']
        
        is_si = ET.SubElement(unit_elem, f'{{{namespace}}}isSI')
        is_si.text = unit['isSI']
        
        category = ET.SubElement(unit_elem, f'{{{namespace}}}category')
        category.text = unit['category']
        
        if unit['isBase']:
            ET.SubElement(unit_elem, f'{{{namespace}}}isBase')
        else:
            base_unit = ET.SubElement(unit_elem, f'{{{namespace}}}baseUnit')
            base_unit.text = unit['baseUnit']
            
            conv_ref = ET.SubElement(unit_elem, f'{{{namespace}}}conversionRef')
            conv_ref.text = unit['conversionRef']
            
            is_exact = ET.SubElement(unit_elem, f'{{{namespace}}}isExact')
            is_exact.text = unit['isExact']
            
            a = ET.SubElement(unit_elem, f'{{{namespace}}}A')
            a.text = unit['A']
            
            b = ET.SubElement(unit_elem, f'{{{namespace}}}B')
            b.text = unit['B']
            
            c = ET.SubElement(unit_elem, f'{{{namespace}}}C')
            c.text = unit['C']
            
            d = ET.SubElement(unit_elem, f'{{{namespace}}}D')
            d.text = unit['D']
        
        if unit['underlyingDef']:
            underlying = ET.SubElement(unit_elem, f'{{{namespace}}}underlyingDef')
            underlying.text = unit['underlyingDef']
        
        if unit['description']:
            desc = ET.SubElement(unit_elem, f'{{{namespace}}}description')
            desc.text = unit['description']
    
    # Reference Set
    ref_set = ET.SubElement(root, f'{{{namespace}}}referenceSet', version=data['version'])
    for ref in data['references']:
        ref_elem = ET.SubElement(ref_set, f'{{{namespace}}}reference')
        
        ref_id = ET.SubElement(ref_elem, f'{{{namespace}}}ID')
        ref_id.text = ref['ID']
        
        desc = ET.SubElement(ref_elem, f'{{{namespace}}}description')
        desc.text = ref['description']
    
    # Prefix Set
    prefix_set = ET.SubElement(root, f'{{{namespace}}}prefixSet', version=data['version'])
    for prefix in data['prefixes']:
        prefix_elem = ET.SubElement(prefix_set, f'{{{namespace}}}prefix')
        
        symbol = ET.SubElement(prefix_elem, f'{{{namespace}}}symbol')
        symbol.text = prefix['symbol']
        
        name = ET.SubElement(prefix_elem, f'{{{namespace}}}name')
        name.text = prefix['name']
        
        multiplier = ET.SubElement(prefix_elem, f'{{{namespace}}}multiplier')
        multiplier.text = prefix['multiplier']
        
        if prefix['commonName']:
            common = ET.SubElement(prefix_elem, f'{{{namespace}}}commonName')
            common.text = prefix['commonName']
    
    return root


def prettify_xml(elem):
    """Return a pretty-printed XML string."""
    rough_string = ET.tostring(elem, encoding='unicode')
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent='  ')


def main():
    """Main entry point."""
    
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    
    excel_file = sys.argv[1]
    output_file = sys.argv[2]
    namespace = sys.argv[3] if len(sys.argv) > 3 else 'http://www.energistics.org/energyml/data/uomv1'
    
    print(f"Reading Excel: {excel_file}")
    data = read_excel_data(excel_file)
    
    print(f"Dictionary version: {data['version']}")
    if data.get('title'):
        print(f"Title: {data['title']}")
    if data.get('namespace'):
        print(f"Namespace: {data['namespace']}")
    if data.get('schemaLocation'):
        print(f"Schema Location: {data['schemaLocation'][:80]}...")
    
    print(f"\nFound:")
    print(f"  - {len(data['unit_dimensions'])} unit dimensions")
    print(f"  - {len(data['quantity_classes'])} quantity classes")
    print(f"  - {len(data['units'])} units")
    print(f"  - {len(data['references'])} references")
    print(f"  - {len(data['prefixes'])} prefixes")
    
    print(f"\nGenerating XML...")
    # Use namespace from command line if provided, otherwise use one from Excel
    root = create_xml(data, namespace if len(sys.argv) > 3 else None)
    
    print(f"Writing to: {output_file}")
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(prettify_xml(root))
    
    print("Done!")


if __name__ == '__main__':
    main()
