#!/usr/bin/env python3
"""
Energistics UOM Dictionary XML to Excel Converter

Reads an Energistics Unit of Measure Dictionary XML file and populates
an Excel workbook template with the data for easier inspection and editing.

Usage:
    python uom_xml_to_excel.py <input_xml> <template_xlsx> <output_xlsx>
    
Example:
    python uom_xml_to_excel.py uom_dictionary.xml uom_dictionary_template.xlsx uom_output.xlsx
"""

import sys
import xml.etree.ElementTree as ET
from openpyxl import load_workbook


def parse_uom_dictionary(xml_file):
    """Parse the UOM dictionary XML file and extract data structures."""
    
    tree = ET.parse(xml_file)
    root = tree.getroot()
    
    # Extract namespace from root tag
    namespace = ''
    if '}' in root.tag:
        namespace = root.tag.split('}')[0].strip('{')
    
    # Create namespace dictionary for find operations
    ns = {'ns': namespace} if namespace else {}
    
    def get_text(element, tag_name, default=''):
        """Safely get text from an XML element by tag name."""
        if namespace:
            found = element.find(f'ns:{tag_name}', ns)
        else:
            found = element.find(tag_name)
        return found.text if found is not None and found.text else default
    
    def find_all(element, tag_name):
        """Find all elements matching tag name."""
        if namespace:
            return element.findall(f'ns:{tag_name}', ns)
        else:
            return element.findall(tag_name)
    
    def find_one(element, tag_name):
        """Find one element matching tag name."""
        if namespace:
            return element.find(f'ns:{tag_name}', ns)
        else:
            return element.find(tag_name)
    
    # Find root dictionary element
    if root.tag.endswith('uomDictionary') or root.tag == 'uomDictionary':
        dict_elem = root
    else:
        raise ValueError("Root element is not uomDictionary")
    
    # Extract schema location attribute
    xsi_ns = 'http://www.w3.org/2001/XMLSchema-instance'
    schema_location = dict_elem.get(f'{{{xsi_ns}}}schemaLocation', '')
    
    data = {
        'version': dict_elem.get('version', ''),
        'namespace': namespace,
        'schemaLocation': schema_location,
        'title': get_text(dict_elem, 'title'),
        'originator': get_text(dict_elem, 'originator'),
        'description': get_text(dict_elem, 'description'),
        'unit_dimensions': [],
        'quantity_classes': [],
        'units': [],
        'references': [],
        'prefixes': []
    }
    
    # Parse Unit Dimensions
    unit_dim_set = find_one(dict_elem, 'unitDimensionSet')
    if unit_dim_set is not None:
        for ud in find_all(unit_dim_set, 'unitDimension'):
            data['unit_dimensions'].append({
                'name': get_text(ud, 'name'),
                'dimension': get_text(ud, 'dimension'),
                'baseForConversion': get_text(ud, 'baseForConversion'),
                'canonicalUnit': get_text(ud, 'canonicalUnit'),
                'description': get_text(ud, 'description')
            })
    
    # Parse Quantity Classes
    qty_class_set = find_one(dict_elem, 'quantityClassSet')
    if qty_class_set is not None:
        for qc in find_all(qty_class_set, 'quantityClass'):
            member_units = [mu.text for mu in find_all(qc, 'memberUnit') if mu.text]
            data['quantity_classes'].append({
                'name': get_text(qc, 'name'),
                'dimension': get_text(qc, 'dimension'),
                'baseForConversion': get_text(qc, 'baseForConversion'),
                'alternativeBase': get_text(qc, 'alternativeBase'),
                'description': get_text(qc, 'description'),
                'memberUnits': ', '.join(member_units)
            })
    
    # Parse Units
    unit_set = find_one(dict_elem, 'unitSet')
    if unit_set is not None:
        for unit in find_all(unit_set, 'unit'):
            is_base_elem = find_one(unit, 'isBase')
            is_base = is_base_elem is not None
            
            data['units'].append({
                'symbol': get_text(unit, 'symbol'),
                'name': get_text(unit, 'name'),
                'dimension': get_text(unit, 'dimension'),
                'isSI': get_text(unit, 'isSI'),
                'category': get_text(unit, 'category'),
                'isBase': 'TRUE' if is_base else 'FALSE',
                'baseUnit': '' if is_base else get_text(unit, 'baseUnit'),
                'conversionRef': '' if is_base else get_text(unit, 'conversionRef'),
                'isExact': '' if is_base else get_text(unit, 'isExact'),
                'A': '' if is_base else get_text(unit, 'A'),
                'B': '' if is_base else get_text(unit, 'B'),
                'C': '' if is_base else get_text(unit, 'C'),
                'D': '' if is_base else get_text(unit, 'D'),
                'underlyingDef': get_text(unit, 'underlyingDef'),
                'description': get_text(unit, 'description')
            })
    
    # Parse References
    ref_set = find_one(dict_elem, 'referenceSet')
    if ref_set is not None:
        for ref in find_all(ref_set, 'reference'):
            data['references'].append({
                'ID': get_text(ref, 'ID'),
                'description': get_text(ref, 'description')
            })
    
    # Parse Prefixes
    prefix_set = find_one(dict_elem, 'prefixSet')
    if prefix_set is not None:
        for prefix in find_all(prefix_set, 'prefix'):
            data['prefixes'].append({
                'symbol': get_text(prefix, 'symbol'),
                'name': get_text(prefix, 'name'),
                'multiplier': get_text(prefix, 'multiplier'),
                'commonName': get_text(prefix, 'commonName')
            })
    
    return data


def populate_excel(template_file, output_file, data):
    """Populate the Excel template with data from the UOM dictionary."""
    
    wb = load_workbook(template_file)
    
    # Populate Metadata
    if 'Metadata' in wb.sheetnames:
        metadata = wb['Metadata']
        metadata['B1'] = data['version']
        
        # Add namespace and schema location
        row = 3
        if data.get('namespace'):
            metadata.cell(row, 1, 'Namespace')
            metadata.cell(row, 2, data['namespace'])
            row += 1
        if data.get('schemaLocation'):
            metadata.cell(row, 1, 'Schema Location')
            metadata.cell(row, 2, data['schemaLocation'])
            metadata.column_dimensions['B'].width = 100
            row += 1
        
        # Add citation info
        if data.get('title'):
            metadata.cell(row, 1, 'Title')
            metadata.cell(row, 2, data['title'])
            row += 1
        if data.get('originator'):
            metadata.cell(row, 1, 'Originator')
            metadata.cell(row, 2, data['originator'])
            row += 1
        if data.get('description'):
            metadata.cell(row, 1, 'Description')
            metadata.cell(row, 2, data['description'])
            metadata.row_dimensions[row].height = 100
            from openpyxl.styles import Alignment
            metadata.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='top')
            row += 1
    
    # Populate Unit Dimensions
    if 'Unit Dimensions' in wb.sheetnames:
        sheet = wb['Unit Dimensions']
        row = 2
        for ud in data['unit_dimensions']:
            sheet.cell(row, 1, ud['name'])
            sheet.cell(row, 2, ud['dimension'])
            sheet.cell(row, 3, ud['baseForConversion'])
            sheet.cell(row, 4, ud['canonicalUnit'])
            sheet.cell(row, 5, ud['description'])
            row += 1
    
    # Populate Quantity Classes
    if 'Quantity Classes' in wb.sheetnames:
        sheet = wb['Quantity Classes']
        row = 2
        for qc in data['quantity_classes']:
            sheet.cell(row, 1, qc['name'])
            sheet.cell(row, 2, qc['dimension'])
            sheet.cell(row, 3, qc['baseForConversion'])
            sheet.cell(row, 4, qc['alternativeBase'])
            sheet.cell(row, 5, qc['description'])
            sheet.cell(row, 6, qc['memberUnits'])
            row += 1
    
    # Populate Units
    if 'Units' in wb.sheetnames:
        sheet = wb['Units']
        row = 3  # Skip note row
        for unit in data['units']:
            sheet.cell(row, 1, unit['symbol'])
            sheet.cell(row, 2, unit['name'])
            sheet.cell(row, 3, unit['dimension'])
            sheet.cell(row, 4, unit['isSI'])
            sheet.cell(row, 5, unit['category'])
            sheet.cell(row, 6, unit['isBase'])
            sheet.cell(row, 7, unit['baseUnit'])
            sheet.cell(row, 8, unit['conversionRef'])
            sheet.cell(row, 9, unit['isExact'])
            sheet.cell(row, 10, unit['A'])
            sheet.cell(row, 11, unit['B'])
            sheet.cell(row, 12, unit['C'])
            sheet.cell(row, 13, unit['D'])
            sheet.cell(row, 14, unit['underlyingDef'])
            sheet.cell(row, 15, unit['description'])
            row += 1
    
    # Populate References
    if 'References' in wb.sheetnames:
        sheet = wb['References']
        row = 2
        for ref in data['references']:
            sheet.cell(row, 1, ref['ID'])
            sheet.cell(row, 2, ref['description'])
            row += 1
    
    # Populate Prefixes
    if 'Prefixes' in wb.sheetnames:
        sheet = wb['Prefixes']
        row = 2
        for prefix in data['prefixes']:
            sheet.cell(row, 1, prefix['symbol'])
            sheet.cell(row, 2, prefix['name'])
            sheet.cell(row, 3, prefix['multiplier'])
            sheet.cell(row, 4, prefix['commonName'])
            row += 1
    
    wb.save(output_file)


def main():
    """Main entry point."""
    
    if len(sys.argv) != 4:
        print(__doc__)
        sys.exit(1)
    
    xml_file = sys.argv[1]
    template_file = sys.argv[2]
    output_file = sys.argv[3]
    
    print(f"Reading XML: {xml_file}")
    data = parse_uom_dictionary(xml_file)
    
    print(f"Dictionary version: {data['version']}")
    if data.get('title'):
        print(f"Title: {data['title']}")
    print(f"\nFound:")
    print(f"  - {len(data['unit_dimensions'])} unit dimensions")
    print(f"  - {len(data['quantity_classes'])} quantity classes")
    print(f"  - {len(data['units'])} units")
    print(f"  - {len(data['references'])} references")
    print(f"  - {len(data['prefixes'])} prefixes")
    
    print(f"\nPopulating template: {template_file}")
    populate_excel(template_file, output_file, data)
    
    print(f"Output saved to: {output_file}")


if __name__ == '__main__':
    main()
